"""
Build Database - Genera todos los formatos de base de datos para Yolitia.

Entrada: scraping_processed.json (158 productos)
Salida:
  - yolitia.db (SQLite - fuente de verdad)
  - yolitia_schema.sql (Schema documentado para IA)
  - yolitia_products_database.json (JSON universal)
  - yolitia_woocommerce_import.csv
  - yolitia_prestashop_import.csv
  - yolitia_medusa_import.json
  - yolitia_products_database.xlsx
"""

import json
import csv
import sqlite3
import re
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


DATA_DIR = Path(__file__).parent.parent / "data"
INPUT_FILE = DATA_DIR / "scraping_processed.json"
DB_FILE = DATA_DIR / "yolitia.db"
SCHEMA_FILE = DATA_DIR / "yolitia_schema.sql"
JSON_FILE = DATA_DIR / "yolitia_products_database.json"
WOO_CSV = DATA_DIR / "yolitia_woocommerce_import.csv"
PRESTA_CSV = DATA_DIR / "yolitia_prestashop_import.csv"
MEDUSA_JSON = DATA_DIR / "yolitia_medusa_import.json"
XLSX_FILE = DATA_DIR / "yolitia_products_database.xlsx"


# ============================================================
# COLORES PARA XLSX
# ============================================================
COLOR_OLIVA = "7A8C6E"
COLOR_MARFIL = "F8F6F0"
COLOR_AZUL = "B8CCE4"
COLOR_BLANCO = "FFFFFF"
COLOR_NEGRO = "1A1A1A"


# ============================================================
# SQL SCHEMA
# ============================================================
SQL_SCHEMA = '''
-- ============================================================
-- YOLITIA - Base de Datos de Productos
-- Catalogo Digital + E-Commerce
-- ============================================================
-- Plataformas soportadas: WooCommerce, PrestaShop, Medusa.js, Magento
-- Legibilidad IA: Comentarios descriptivos, nombres claros, datos de ejemplo
-- ============================================================

-- ============================================================
-- TABLA: categorias
-- Descripcion: Taxonomia de productos en arbol jerarquico
-- Uso: Clasificar productos para navegacion y filtros
-- ============================================================
CREATE TABLE categorias (
    id TEXT PRIMARY KEY,                    -- Formato: "CAT-001"
    nombre TEXT NOT NULL,                   -- Ejemplo: "Joyeria & Accesorios"
    slug TEXT UNIQUE NOT NULL,              -- URL-friendly: "joyeria-accesorios"
    categoria_padre_id TEXT,                -- NULL = categoria raiz
    descripcion TEXT,                       -- Descripcion para SEO
    orden INTEGER DEFAULT 0,                -- Orden de visualizacion
    imagen_url TEXT,                        -- Imagen representativa
    activo INTEGER DEFAULT 1,               -- 1=sí, 0=no
    creado_en TEXT DEFAULT (datetime('now')),
    actualizado_en TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (categoria_padre_id) REFERENCES categorias(id)
);

-- ============================================================
-- TABLA: productos
-- Descripcion: Catalogo maestro de productos Yolitia
-- Uso: Fuente de verdad para catalogo PDF y tiendas online
-- ============================================================
CREATE TABLE productos (
    id TEXT PRIMARY KEY,                    -- Formato: "YOL-001"
    sku TEXT UNIQUE NOT NULL,               -- Formato: "YOL-JAR-001"
    codigo_barras TEXT,                     -- EAN/UPC (para inventario fisico)
    gtin TEXT,                              -- Global Trade Item Number
    mpn TEXT,                               -- Manufacturer Part Number
    
    -- Identificacion
    nombre_original TEXT NOT NULL,          -- Nombre en MakerWorld
    nombre_yolitia TEXT NOT NULL,           -- Nombre creativo en español
    handle TEXT UNIQUE NOT NULL,            -- URL slug: "aretes-lluvia-del-bosque"
    slug TEXT UNIQUE NOT NULL,              -- Alias de handle
    
    -- Clasificacion
    categoria_id TEXT NOT NULL,             -- FK a categorias
    subcategoria_id TEXT,                   -- FK a categorias (hijo)
    tipo_producto TEXT DEFAULT 'simple',    -- simple, variable, agrupado
    
    -- Descripciones (multi-formato)
    descripcion_corta TEXT,                 -- Max 100 chars, para catalogo
    descripcion_larga TEXT,                 -- 2-3 oraciones, para tienda
    descripcion_html TEXT,                  -- HTML para rich content
    
    -- Material y especificaciones
    material TEXT DEFAULT 'PLA',            -- Material principal
    beneficios_material TEXT,               -- JSON array: '["ligero","resistente"]'
    peso_gramos REAL,                       -- Peso del producto
    dimensiones TEXT,                       -- JSON: '{"largo":3,"ancho":2,"alto":1}'
    
    -- Precios (null = pendiente de definir)
    precio REAL,                            -- Precio regular (MXN)
    precio_comparacion REAL,                -- Precio antes de oferta
    precio_costo REAL,                      -- Costo de produccion
    moneda TEXT DEFAULT 'MXN',
    impuesto_porcentaje REAL DEFAULT 16,    -- IVA Mexico
    impuesto_incluido INTEGER DEFAULT 0,    -- 0=precio sin IVA
    
    -- Estado y visibilidad
    activo INTEGER DEFAULT 1,               -- 1=visible, 0=oculto
    estado TEXT DEFAULT 'borrador',         -- borrador, publicado, agotado
    visibilidad TEXT DEFAULT 'visible',     -- visible, catalogo, busqueda, oculto
    destacado INTEGER DEFAULT 0,            -- 1=producto destacado
    
    -- Catalogo PDF
    pagina_catalogo INTEGER,                -- Numero de pagina en PDF
    posicion_pagina INTEGER,                -- Posicion dentro de la pagina
    
    -- Relaciones
    url_makerworld TEXT,                    -- URL al modelo original
    url_externa TEXT,                       -- URL alternativa
    
    -- Metadata para IA
    metadata TEXT,                          -- JSON con datos extendidos
    tags TEXT,                              -- JSON array: '["aretes","joyeria"]'
    
    -- Timestamps
    creado_en TEXT DEFAULT (datetime('now')),
    actualizado_en TEXT DEFAULT (datetime('now')),
    publicado_en TEXT,
    
    FOREIGN KEY (categoria_id) REFERENCES categorias(id),
    FOREIGN KEY (subcategoria_id) REFERENCES categorias(id)
);

-- ============================================================
-- TABLA: variantes
-- Descripcion: Variantes de producto (color, tamaño, etc.)
-- Uso: Productos con opciones multiples
-- ============================================================
CREATE TABLE variantes (
    id TEXT PRIMARY KEY,                    -- Formato: "YOL-001-VRD-01"
    producto_id TEXT NOT NULL,              -- FK a productos
    sku TEXT UNIQUE NOT NULL,               -- SKU unico por variante
    codigo_barras TEXT,
    
    -- Opciones de variante
    opciones TEXT NOT NULL,                 -- JSON: '{"Color":"Verde","Tamaño":"Unico"}'
    
    -- Precios (hereda del producto si es null)
    precio REAL,
    precio_comparacion REAL,
    peso_gramos REAL,
    
    -- Inventario
    stock_cantidad INTEGER DEFAULT 0,
    stock_tracking INTEGER DEFAULT 0,       -- 1=controlar stock
    stock_minimo INTEGER DEFAULT 0,         -- Alerta de stock bajo
    
    -- Imagen especifica de variante
    imagen_id TEXT,                         -- FK a imagenes (si aplica)
    
    -- Estado
    activo INTEGER DEFAULT 1,
    
    creado_en TEXT DEFAULT (datetime('now')),
    actualizado_en TEXT DEFAULT (datetime('now')),
    
    FOREIGN KEY (producto_id) REFERENCES productos(id) ON DELETE CASCADE
);

-- ============================================================
-- TABLA: imagenes
-- Descripcion: Galeria de imagenes por producto
-- Uso: Catalogo PDF, tienda online, redes sociales
-- ============================================================
CREATE TABLE imagenes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    producto_id TEXT NOT NULL,              -- FK a productos
    variante_id TEXT,                       -- FK a variantes (si es especifica)
    
    -- Identificacion
    url TEXT NOT NULL,                      -- URL publica de la imagen
    filename TEXT NOT NULL,                 -- Nombre: "YOL-001_principal_01.jpg"
    tipo TEXT NOT NULL,                     -- principal, detalle, ambiente, variante
    
    -- Metadata SEO
    alt_text TEXT,                          -- Texto alternativo para accesibilidad
    titulo TEXT,                            -- Titulo de la imagen
    
    -- Orden y uso
    posicion INTEGER DEFAULT 0,             -- Orden en galeria
    es_principal INTEGER DEFAULT 0,         -- 1=imagen principal del producto
    
    -- Almacenamiento
    ruta_local TEXT,                        -- Ruta en disco local
    drive_url TEXT,                         -- URL en Google Drive
    ancho INTEGER,                          -- Pixeles de ancho
    alto INTEGER,                           -- Pixeles de alto
    tamano_kb INTEGER,                      -- Tamaño del archivo
    
    creado_en TEXT DEFAULT (datetime('now')),
    
    FOREIGN KEY (producto_id) REFERENCES productos(id) ON DELETE CASCADE,
    FOREIGN KEY (variante_id) REFERENCES variantes(id) ON DELETE SET NULL
);

-- ============================================================
-- TABLA: seo
-- Descripcion: Metadata SEO por producto
-- Uso: Optimizacion para buscadores y redes sociales
-- ============================================================
CREATE TABLE seo (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    producto_id TEXT UNIQUE NOT NULL,       -- FK a productos (1:1)
    
    -- Meta tags
    meta_title TEXT,                        -- Max 60 chars
    meta_description TEXT,                  -- Max 160 chars
    meta_keywords TEXT,                     -- JSON array de palabras clave
    
    -- Open Graph (redes sociales)
    og_title TEXT,
    og_description TEXT,
    og_image TEXT,                          -- URL imagen para compartir
    
    -- Google Shopping
    google_product_category TEXT,           -- Categoria Google Merchant
    condition TEXT DEFAULT 'new',           -- new, refurbished, used
    gender TEXT,                            -- male, female, unisex
    age_group TEXT,                         -- adult, kids
    
    -- Canonical
    canonical_url TEXT,
    
    actualizado_en TEXT DEFAULT (datetime('now')),
    
    FOREIGN KEY (producto_id) REFERENCES productos(id) ON DELETE CASCADE
);

-- ============================================================
-- TABLA: inventario
-- Descripcion: Control de stock y ubicaciones
-- Uso: Gestion de inventario fisico
-- ============================================================
CREATE TABLE inventario (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    producto_id TEXT NOT NULL,              -- FK a productos
    variante_id TEXT,                       -- FK a variantes (si aplica)
    
    -- Stock
    cantidad INTEGER DEFAULT 0,
    reservado INTEGER DEFAULT 0,            -- En ordenes pendientes
    disponible INTEGER GENERATED ALWAYS AS (cantidad - reservado) STORED,
    
    -- Ubicacion
    ubicacion TEXT,                         -- Ejemplo: "Almacen A - Estante 3"
    codigo_ubicacion TEXT,
    
    -- Configuracion
    tracking_activo INTEGER DEFAULT 0,      -- 1=controlar stock
    backorders_permitidos INTEGER DEFAULT 0,
    cantidad_minima INTEGER DEFAULT 0,      -- Punto de reorden
    
    -- Timestamps
    actualizado_en TEXT DEFAULT (datetime('now')),
    
    FOREIGN KEY (producto_id) REFERENCES productos(id) ON DELETE CASCADE,
    FOREIGN KEY (variante_id) REFERENCES variantes(id) ON DELETE SET NULL
);

-- ============================================================
-- TABLA: envio
-- Descripcion: Configuracion de envio por producto
-- Uso: Calculo de costos de envio
-- ============================================================
CREATE TABLE envio (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    producto_id TEXT UNIQUE NOT NULL,       -- FK a productos (1:1)
    
    -- Dimensiones de empaque
    peso_kg REAL,                           -- Peso con empaque
    largo_cm REAL,
    ancho_cm REAL,
    alto_cm REAL,
    
    -- Clasificacion
    clase_envio TEXT DEFAULT 'estandar',    -- estandar, express, economico
    requiere_firma INTEGER DEFAULT 0,
    es_fragil INTEGER DEFAULT 0,
    
    -- Origen
    codigo_postal_origen TEXT,
    pais_origen TEXT DEFAULT 'MX',
    
    -- Restricciones
    requiere_envio INTEGER DEFAULT 1,       -- 1=requiere envio, 0=digital/pickup
    gravable INTEGER DEFAULT 1,             -- 1=paga impuestos
    
    actualizado_en TEXT DEFAULT (datetime('now')),
    
    FOREIGN KEY (producto_id) REFERENCES productos(id) ON DELETE CASCADE
);

-- ============================================================
-- TABLA: productos_relacionados
-- Descripcion: Relaciones entre productos (upsell, cross-sell)
-- Uso: Recomendaciones y ventas cruzadas
-- ============================================================
CREATE TABLE productos_relacionados (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    producto_id TEXT NOT NULL,              -- Producto principal
    producto_relacionado_id TEXT NOT NULL,  -- Producto relacionado
    tipo TEXT NOT NULL,                     -- upsell, cross_sell, agrupado
    orden INTEGER DEFAULT 0,
    
    FOREIGN KEY (producto_id) REFERENCES productos(id) ON DELETE CASCADE,
    FOREIGN KEY (producto_relacionado_id) REFERENCES productos(id) ON DELETE CASCADE,
    UNIQUE(producto_id, producto_relacionado_id, tipo)
);

-- ============================================================
-- TABLA: colecciones
-- Descripcion: Agrupaciones de productos (marketing)
-- Uso: Colecciones tematicas, temporadas, promociones
-- ============================================================
CREATE TABLE colecciones (
    id TEXT PRIMARY KEY,                    -- Formato: "COL-001"
    nombre TEXT NOT NULL,
    slug TEXT UNIQUE NOT NULL,
    descripcion TEXT,
    imagen_url TEXT,
    
    -- Configuracion
    tipo TEXT DEFAULT 'manual',             -- manual, automatica
    condiciones TEXT,                       -- JSON para colecciones automaticas
    orden INTEGER DEFAULT 0,
    
    -- Estado
    activo INTEGER DEFAULT 1,
    destacada INTEGER DEFAULT 0,
    
    -- Fechas
    fecha_inicio TEXT,                      -- Para colecciones temporales
    fecha_fin TEXT,
    
    creado_en TEXT DEFAULT (datetime('now')),
    actualizado_en TEXT DEFAULT (datetime('now'))
);

-- ============================================================
-- TABLA: colecciones_productos
-- Descripcion: Relacion muchos-a-muchos entre colecciones y productos
-- ============================================================
CREATE TABLE colecciones_productos (
    coleccion_id TEXT NOT NULL,
    producto_id TEXT NOT NULL,
    orden INTEGER DEFAULT 0,
    
    PRIMARY KEY (coleccion_id, producto_id),
    FOREIGN KEY (coleccion_id) REFERENCES colecciones(id) ON DELETE CASCADE,
    FOREIGN KEY (producto_id) REFERENCES productos(id) ON DELETE CASCADE
);

-- ============================================================
-- TABLA: platform_data
-- Descripcion: Datos especificos por plataforma e-commerce
-- Uso: Mapeo de IDs y configuraciones por plataforma
-- ============================================================
CREATE TABLE platform_data (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    producto_id TEXT NOT NULL,              -- FK a productos
    variante_id TEXT,                       -- FK a variantes (si aplica)
    
    plataforma TEXT NOT NULL,               -- woocommerce, prestashop, medusa, magento
    platform_id TEXT,                       -- ID en la plataforma externa
    platform_sku TEXT,                      -- SKU en la plataforma
    
    -- Datos extendidos
    datos TEXT,                             -- JSON con datos especificos
    
    -- Sync
    ultimo_sync TEXT,
    estado_sync TEXT DEFAULT 'pendiente',   -- pendiente, sincronizado, error
    
    FOREIGN KEY (producto_id) REFERENCES productos(id) ON DELETE CASCADE,
    FOREIGN KEY (variante_id) REFERENCES variantes(id) ON DELETE SET NULL,
    UNIQUE(producto_id, variante_id, plataforma)
);

-- ============================================================
-- INDICES para rendimiento
-- ============================================================
CREATE INDEX idx_productos_categoria ON productos(categoria_id);
CREATE INDEX idx_productos_estado ON productos(estado);
CREATE INDEX idx_productos_activo ON productos(activo);
CREATE INDEX idx_productos_destacado ON productos(destacado);
CREATE INDEX idx_variantes_producto ON variantes(producto_id);
CREATE INDEX idx_imagenes_producto ON imagenes(producto_id);
CREATE INDEX idx_imagenes_principal ON imagenes(producto_id, es_principal);
CREATE INDEX idx_inventario_producto ON inventario(producto_id);
CREATE INDEX idx_colecciones_productos_producto ON colecciones_productos(producto_id);

-- ============================================================
-- VISTAS para consultas comunes
-- ============================================================

-- Vista: Productos completos con categoria e imagen principal
CREATE VIEW vista_productos_completos AS
SELECT 
    p.*,
    c.nombre as categoria_nombre,
    c.slug as categoria_slug,
    i.url as imagen_principal_url,
    i.alt_text as imagen_principal_alt,
    s.meta_title,
    s.meta_description,
    COALESCE(inv.cantidad, 0) as stock_total
FROM productos p
LEFT JOIN categorias c ON p.categoria_id = c.id
LEFT JOIN imagenes i ON p.id = i.producto_id AND i.es_principal = 1
LEFT JOIN seo s ON p.id = s.producto_id
LEFT JOIN inventario inv ON p.id = inv.producto_id AND inv.variante_id IS NULL;

-- Vista: Productos para exportacion e-commerce
CREATE VIEW vista_productos_ecommerce AS
SELECT 
    p.id,
    p.sku,
    p.nombre_yolitia as nombre,
    p.handle,
    p.descripcion_corta,
    p.descripcion_larga,
    p.precio,
    p.precio_comparacion,
    p.peso_gramos,
    c.nombre as categoria,
    c.slug as categoria_slug,
    p.tags,
    p.activo,
    p.estado,
    GROUP_CONCAT(DISTINCT i.url) as imagenes_urls
FROM productos p
LEFT JOIN categorias c ON p.categoria_id = c.id
LEFT JOIN imagenes i ON p.id = i.producto_id
WHERE p.activo = 1
GROUP BY p.id;
'''


def load_input():
    with open(INPUT_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    seen_handles = {}
    for p in data["productos"]:
        handle = p["handle"]
        if handle in seen_handles:
            p["handle"] = f"{handle}-{p['id'].lower()}"
        else:
            seen_handles[handle] = True
    
    return data


def build_categories(productos):
    categories = {}
    cat_id = 1
    subcat_id = 100
    
    cat_parent_map = {
        "Joyería & Accesorios": None,
        "Hogar & Decoración": None,
        "Organización": None,
        "Regalos & Coleccionables": None,
        "Entretenimiento": None,
    }
    
    subcat_parent_map = {
        "Aretes": "Joyería & Accesorios",
        "Collares": "Joyería & Accesorios",
        "Pulseras": "Joyería & Accesorios",
        "Anillos": "Joyería & Accesorios",
        "Macetas": "Hogar & Decoración",
        "Iluminación": "Hogar & Decoración",
        "Relojes": "Hogar & Decoración",
        "Figuras Decorativas": "Hogar & Decoración",
        "Decoración": "Hogar & Decoración",
        "Joyeros": "Organización",
        "Porta-Objetos": "Organización",
        "Organizadores": "Organización",
        "Llaveros": "Regalos & Coleccionables",
        "Figuras Temáticas": "Regalos & Coleccionables",
        "Coleccionables": "Regalos & Coleccionables",
        "Fidgets": "Entretenimiento",
        "Rompecabezas": "Entretenimiento",
        "Juguetes": "Entretenimiento",
    }
    
    for cat_name, parent in cat_parent_map.items():
        cat_id_str = f"CAT-{cat_id:03d}"
        slug = cat_name.lower().replace(" & ", "-").replace(" ", "-")
        categories[cat_name] = {
            "id": cat_id_str,
            "nombre": cat_name,
            "slug": slug,
            "parent_id": None,
            "orden": cat_id
        }
        cat_id += 1
    
    for subcat_name, parent_name in subcat_parent_map.items():
        subcat_id_str = f"CAT-{subcat_id:03d}"
        slug = subcat_name.lower().replace(" ", "-")
        parent_id = categories.get(parent_name, {}).get("id")
        categories[subcat_name] = {
            "id": subcat_id_str,
            "nombre": subcat_name,
            "slug": slug,
            "parent_id": parent_id,
            "orden": subcat_id - 99
        }
        subcat_id += 1
    
    return categories


def build_sqlite(productos, categories):
    if DB_FILE.exists():
        DB_FILE.unlink()
    
    conn = sqlite3.connect(str(DB_FILE))
    conn.execute("PRAGMA foreign_keys = ON")
    cursor = conn.cursor()
    
    cursor.executescript(SQL_SCHEMA)
    
    for cat_name, cat_data in categories.items():
        cursor.execute(
            "INSERT INTO categorias (id, nombre, slug, categoria_padre_id, orden, activo) VALUES (?, ?, ?, ?, ?, 1)",
            (cat_data["id"], cat_data["nombre"], cat_data["slug"], cat_data["parent_id"], cat_data["orden"])
        )
    
    for p in productos:
        cat_id = categories.get(p["categoria"], {}).get("id", "CAT-001")
        subcat_id = categories.get(p["subcategoria"], {}).get("id")
        
        handle = p["handle"]
        slug = handle
        
        cursor.execute("SELECT id FROM productos WHERE handle = ? OR slug = ?", (handle, slug))
        if cursor.fetchone():
            handle = f"{handle}-{p['id'].lower()}"
            slug = handle
        
        cursor.execute(
            """INSERT INTO productos (
                id, sku, nombre_original, nombre_yolitia, handle, slug,
                categoria_id, subcategoria_id, tipo_producto,
                descripcion_corta, descripcion_larga, descripcion_html,
                material, beneficios_material,
                precio, moneda, impuesto_porcentaje,
                activo, estado, visibilidad, destacado,
                url_makerworld, tags, metadata,
                pagina_catalogo, posicion_pagina
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                p["id"], p["sku"], p["nombre_original"], p["nombre_yolitia"],
                handle, slug,
                cat_id, subcat_id, p["tipo_producto"],
                p.get("descripcion_corta"), p.get("descripcion_larga"), p.get("descripcion_html"),
                p["material"], json.dumps(p.get("beneficios_material", []), ensure_ascii=False),
                p.get("precio"), p["moneda"], p.get("impuesto_porcentaje", 16),
                1 if p["activo"] else 0, p["estado"], p["visibilidad"],
                1 if p["catalogo"].get("destacado") else 0,
                p["url_makerworld"],
                json.dumps(p.get("seo", {}).get("meta_keywords", []), ensure_ascii=False),
                json.dumps(p.get("metadata", {}), ensure_ascii=False),
                p["catalogo"].get("pagina"), p["catalogo"].get("posicion_pagina")
            )
        )
        
        for img in p.get("imagenes", []):
            cursor.execute(
                """INSERT INTO imagenes (producto_id, url, filename, tipo, alt_text, titulo, posicion, es_principal, ruta_local, drive_url)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    p["id"], img["url"], img["filename"], img["tipo"],
                    img.get("alt_text"), img.get("titulo"),
                    img.get("posicion", 0), 1 if img.get("es_principal") else 0,
                    img.get("ruta_local"), img.get("drive_url")
                )
            )
        
        cursor.execute(
            """INSERT INTO seo (producto_id, meta_keywords, condition)
               VALUES (?, ?, ?)""",
            (
                p["id"],
                json.dumps(p.get("seo", {}).get("meta_keywords", []), ensure_ascii=False),
                p.get("seo", {}).get("condition", "new")
            )
        )
        
        cursor.execute(
            """INSERT INTO inventario (producto_id, tracking_activo)
               VALUES (?, ?)""",
            (p["id"], 1 if p.get("inventario", {}).get("tracking") else 0)
        )
        
        cursor.execute(
            """INSERT INTO envio (producto_id, clase_envio, requiere_envio, gravable)
               VALUES (?, ?, ?, ?)""",
            (
                p["id"],
                p.get("envio", {}).get("clase_envio", "estandar"),
                1 if p.get("envio", {}).get("requiere_envio", True) else 0,
                1 if p.get("envio", {}).get("gravable", True) else 0
            )
        )
    
    conn.commit()
    conn.close()
    print(f"  SQLite: {DB_FILE}")


def save_schema_sql():
    with open(SCHEMA_FILE, 'w', encoding='utf-8') as f:
        f.write(SQL_SCHEMA)
    print(f"  SQL Schema: {SCHEMA_FILE}")


def build_json_universal(productos, categories):
    output = {
        "metadata": {
            "version": "1.0",
            "fecha_creacion": datetime.now().isoformat(),
            "total_productos": len(productos),
            "moneda": "MXN",
            "idioma": "es-MX",
            "plataformas_soportadas": ["woocommerce", "prestashop", "medusa", "magento"]
        },
        "categorias": list(categories.values()),
        "productos": productos
    }
    
    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"  JSON Universal: {JSON_FILE}")


def build_woocommerce_csv(productos, categories):
    rows = []
    
    for p in productos:
        cat_path = p["categoria"]
        if p["subcategoria"]:
            cat_path = f"{p['categoria']} > {p['subcategoria']}"
        
        images = [img["url"] for img in p.get("imagenes", [])]
        
        row = {
            "ID": "",
            "Type": p["tipo_producto"],
            "SKU": p["sku"],
            "Name": p["nombre_yolitia"],
            "Published": 1 if p["estado"] == "publicado" else 0,
            "Is featured?": 1 if p["catalogo"].get("destacado") else 0,
            "Visibility in catalog": p["visibilidad"],
            "Short description": p.get("descripcion_corta", ""),
            "Description": p.get("descripcion_larga", ""),
            "Date sale price starts": "",
            "Date sale price ends": "",
            "Tax status": "taxable",
            "Tax class": "standard",
            "In stock?": 1,
            "Stock": p.get("inventario", {}).get("cantidad", ""),
            "Low stock amount": "",
            "Backorders allowed?": 0,
            "Sold individually?": 0,
            "Weight (unit)": p.get("peso_gramos", "") or "",
            "Length (unit)": "",
            "Width (unit)": "",
            "Height (unit)": "",
            "Allow customer reviews?": 1,
            "Purchase Note": "",
            "Sale price": p.get("precio_comparacion", "") or "",
            "Regular price": p.get("precio", "") or "",
            "Categories": cat_path,
            "Tags": ",".join(p.get("seo", {}).get("meta_keywords", [])),
            "Shipping class": p.get("envio", {}).get("clase_envio", ""),
            "Images": ",".join(images),
            "Download limit": "",
            "Download expiry days": "",
            "Parent": "",
            "Grouped products": "",
            "Upsells": "",
            "Cross-sells": "",
            "External URL": p.get("url_makerworld", ""),
            "Button text": "",
            "Position": p["catalogo"].get("orden_coleccion", 0),
            "Attribute 1 name": "Material",
            "Attribute 1 value(s)": p["material"],
            "Attribute 1 default": p["material"],
            "Attribute 1 visible": 1,
            "Attribute 1 global": 0
        }
        rows.append(row)
    
    fieldnames = list(rows[0].keys()) if rows else []
    
    with open(WOO_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    
    print(f"  WooCommerce CSV: {WOO_CSV} ({len(rows)} productos)")


def build_prestashop_csv(productos, categories):
    rows = []
    
    for p in productos:
        cat_name = p["subcategoria"] if p["subcategoria"] else p["categoria"]
        
        images = [img["url"] for img in p.get("imagenes", [])]
        
        row = {
            "ID": "",
            "Active (0/1)": 1 if p["activo"] else 0,
            "Name *": p["nombre_yolitia"],
            "Categories (x,y,z...)": cat_name,
            "Price tax excluded": p.get("precio", "") or "",
            "Tax rules id": 1,
            "Wholesale price": p.get("precio_costo", "") or "",
            "On sale (0/1)": 0,
            "Discount amount": p.get("precio_comparacion", "") or "",
            "Discount percent": "",
            "Discount from (yyyy-mm-dd)": "",
            "Discount to (yyyy-mm-dd)": "",
            "Reference #": p["sku"],
            "Supplier reference": "",
            "Supplier": "",
            "Brand": "Yolitia",
            "EAN13": "",
            "UPC": "",
            "Ecotax": "",
            "Weight": p.get("peso_gramos", 0) or 0,
            "Width": "",
            "Height": "",
            "Depth": "",
            "Additional shipping cost": "",
            "Unit price": "",
            "Summary": p.get("descripcion_corta", ""),
            "Description": p.get("descripcion_larga", ""),
            "Tags (x,y,z...)": ",".join(p.get("seo", {}).get("meta_keywords", [])),
            "Short description": p.get("descripcion_corta", ""),
            "Image URLs (x,y,z...)": ",".join(images),
            "Image alt texts (x,y,z...)": ",".join([img.get("alt_text", "") for img in p.get("imagenes", [])]),
            "Visibility": p["visibilidad"],
            "Delete existing images (0/1)": 0,
            "Feature (Name:Value:Position)": f"Material:{p['material']}:1",
            "Available for order (0/1)": 1,
            "Online only (0/1)": 1,
            "Condition": "new",
            "Customizable (0/1)": 0,
            "Uploadable files (0/1)": 0,
            "Text fields (0/1)": 0,
            "Out of stock action": 0,
            "Virtual URL": p["handle"],
            "URL rewrite": p["handle"],
            "Meta title": p.get("seo", {}).get("meta_title", ""),
            "Meta description": p.get("seo", {}).get("meta_description", ""),
            "Meta keywords": ",".join(p.get("seo", {}).get("meta_keywords", [])),
            "ID / Name of shop": ""
        }
        rows.append(row)
    
    fieldnames = list(rows[0].keys()) if rows else []
    
    with open(PRESTA_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    
    print(f"  PrestaShop CSV: {PRESTA_CSV} ({len(rows)} productos)")


def build_medusa_json(productos, categories):
    medusa_products = []
    
    for p in productos:
        cat_name = p["subcategoria"] if p["subcategoria"] else p["categoria"]
        
        images = [{"url": img["url"], "alt_text": img.get("alt_text", "")} for img in p.get("imagenes", [])]
        
        medusa_product = {
            "title": p["nombre_yolitia"],
            "handle": p["handle"],
            "subtitle": p.get("descripcion_corta", ""),
            "description": p.get("descripcion_larga", ""),
            "is_giftcard": False,
            "status": "draft" if p["estado"] == "borrador" else "published",
            "thumbnail": images[0]["url"] if images else None,
            "material": p["material"],
            "weight": p.get("peso_gramos"),
            "length": None,
            "width": None,
            "height": None,
            "hs_code": None,
            "origin_country": "MX",
            "mid_code": None,
            "metadata": {
                "url_makerworld": p["url_makerworld"],
                "tags": p.get("seo", {}).get("meta_keywords", [])
            },
            "options": [
                {
                    "title": "Material",
                    "values": [p["material"]]
                }
            ],
            "variants": [
                {
                    "title": p["nombre_yolitia"],
                    "sku": p["sku"],
                    "barcode": None,
                    "ean": None,
                    "upc": None,
                    "inventory_quantity": p.get("inventario", {}).get("cantidad") or 0,
                    "allow_backorder": p.get("inventario", {}).get("backorders", False),
                    "manage_inventory": p.get("inventario", {}).get("tracking", False),
                    "weight": p.get("peso_gramos"),
                    "length": None,
                    "width": None,
                    "height": None,
                    "hs_code": None,
                    "origin_country": "MX",
                    "mid_code": None,
                    "material": p["material"],
                    "metadata": {},
                    "prices": [
                        {
                            "currency_code": "mxn",
                            "amount": int((p.get("precio") or 0) * 100) if p.get("precio") else None
                        }
                    ],
                    "options": {
                        "Material": p["material"]
                    }
                }
            ],
            "images": images,
            "collection": {
                "title": cat_name,
                "handle": cat_name.lower().replace(" ", "-")
            },
            "tags": [
                {"value": tag} for tag in p.get("seo", {}).get("meta_keywords", [])
            ],
            "sales_channels": [],
            "type": {
                "value": p["subcategoria"] if p["subcategoria"] else p["categoria"]
            }
        }
        
        medusa_products.append(medusa_product)
    
    output = {
        "products": medusa_products
    }
    
    with open(MEDUSA_JSON, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"  Medusa.js JSON: {MEDUSA_JSON} ({len(medusa_products)} productos)")


def build_xlsx(productos, categories):
    wb = openpyxl.Workbook()
    
    header_fill = PatternFill(start_color=COLOR_OLIVA, end_color=COLOR_OLIVA, fill_type="solid")
    header_font = Font(color=COLOR_BLANCO, bold=True, size=11)
    editable_fill = PatternFill(start_color=COLOR_AZUL, end_color=COLOR_AZUL, fill_type="solid")
    row_fill_1 = PatternFill(start_color=COLOR_MARFIL, end_color=COLOR_MARFIL, fill_type="solid")
    row_fill_2 = PatternFill(start_color=COLOR_BLANCO, end_color=COLOR_BLANCO, fill_type="solid")
    border = Border(
        left=Side(style='thin', color='D6D0C8'),
        right=Side(style='thin', color='D6D0C8'),
        top=Side(style='thin', color='D6D0C8'),
        bottom=Side(style='thin', color='D6D0C8')
    )
    
    ws = wb.active
    ws.title = "Productos"
    
    headers = [
        "ID", "SKU", "Nombre Yolitia", "Nombre Original", "Handle",
        "Categoria", "Subcategoria", "Tipo", "Descripcion Corta",
        "Material", "Precio", "Precio Comparacion", "Moneda",
        "Peso (g)", "Activo", "Estado", "URL MakerWorld", "Imagen Principal"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for row_idx, p in enumerate(productos, 2):
        row_fill = row_fill_1 if row_idx % 2 == 0 else row_fill_2
        
        values = [
            p["id"], p["sku"], p["nombre_yolitia"], p["nombre_original"], p["handle"],
            p["categoria"], p["subcategoria"], p["tipo_producto"], p.get("descripcion_corta", ""),
            p["material"],
            None, None, p["moneda"],
            p.get("peso_gramos"),
            "Si" if p["activo"] else "No", p["estado"], p["url_makerworld"],
            p.get("imagen_principal", "")
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            if headers[col-1] in ["Precio", "Precio Comparacion"]:
                cell.fill = editable_fill
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    ws2 = wb.create_sheet("Variantes")
    var_headers = ["Producto ID", "Variante ID", "SKU", "Opciones", "Precio", "Stock", "Activo"]
    for col, header in enumerate(var_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    ws2.cell(row=2, column=1, value="(Sin variantes definidas - productos simples)")
    
    ws3 = wb.create_sheet("Imagenes")
    img_headers = ["Producto ID", "Filename", "Tipo", "URL", "Alt Text", "Posicion", "Es Principal"]
    for col, header in enumerate(img_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    img_row = 2
    for p in productos:
        for img in p.get("imagenes", []):
            values = [
                p["id"], img["filename"], img["tipo"], img["url"],
                img.get("alt_text", ""), img.get("posicion", 0),
                "Si" if img.get("es_principal") else "No"
            ]
            for col, value in enumerate(values, 1):
                cell = ws3.cell(row=img_row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True)
            img_row += 1
    
    ws4 = wb.create_sheet("Categorias")
    cat_headers = ["ID", "Nombre", "Slug", "Categoria Padre", "Orden"]
    for col, header in enumerate(cat_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    cat_row = 2
    for cat_name, cat_data in categories.items():
        values = [cat_data["id"], cat_data["nombre"], cat_data["slug"], cat_data["parent_id"] or "", cat_data["orden"]]
        for col, value in enumerate(values, 1):
            cell = ws4.cell(row=cat_row, column=col, value=value)
            cell.border = border
        cat_row += 1
    
    ws5 = wb.create_sheet("SEO")
    seo_headers = ["Producto ID", "Nombre", "Meta Title", "Meta Description", "Keywords", "Condition"]
    for col, header in enumerate(seo_headers, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for row_idx, p in enumerate(productos, 2):
        keywords = ", ".join(p.get("seo", {}).get("meta_keywords", []))
        values = [
            p["id"], p["nombre_yolitia"],
            p.get("seo", {}).get("meta_title", ""),
            p.get("seo", {}).get("meta_description", ""),
            keywords,
            p.get("seo", {}).get("condition", "new")
        ]
        for col, value in enumerate(values, 1):
            cell = ws5.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)
    
    ws6 = wb.create_sheet("Inventario")
    inv_headers = ["Producto ID", "SKU", "Nombre", "Stock", "Tracking", "Ubicacion", "Backorders"]
    for col, header in enumerate(inv_headers, 1):
        cell = ws6.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for row_idx, p in enumerate(productos, 2):
        values = [
            p["id"], p["sku"], p["nombre_yolitia"],
            None, "No" if not p.get("inventario", {}).get("tracking") else "Si",
            p.get("inventario", {}).get("ubicacion", ""),
            "Si" if p.get("inventario", {}).get("backorders") else "No"
        ]
        for col, value in enumerate(values, 1):
            cell = ws6.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if headers[col-1] if col <= len(headers) else "" == "Stock":
                cell.fill = editable_fill
    
    ws7 = wb.create_sheet("Envio")
    env_headers = ["Producto ID", "SKU", "Peso (kg)", "Largo (cm)", "Ancho (cm)", "Alto (cm)", "Clase", "Requiere Envio"]
    for col, header in enumerate(env_headers, 1):
        cell = ws7.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for row_idx, p in enumerate(productos, 2):
        values = [
            p["id"], p["sku"],
            p.get("envio", {}).get("peso_envio_kg", ""),
            p.get("envio", {}).get("dimensiones_envio_cm", {}).get("largo", "") if p.get("envio", {}).get("dimensiones_envio_cm") else "",
            p.get("envio", {}).get("dimensiones_envio_cm", {}).get("ancho", "") if p.get("envio", {}).get("dimensiones_envio_cm") else "",
            p.get("envio", {}).get("dimensiones_envio_cm", {}).get("alto", "") if p.get("envio", {}).get("dimensiones_envio_cm") else "",
            p.get("envio", {}).get("clase_envio", "estandar"),
            "Si" if p.get("envio", {}).get("requiere_envio", True) else "No"
        ]
        for col, value in enumerate(values, 1):
            cell = ws7.cell(row=row_idx, column=col, value=value)
            cell.border = border
    
    ws8 = wb.create_sheet("Mapeo")
    map_headers = ["Campo Universal", "WooCommerce", "PrestaShop", "Medusa.js", "Magento"]
    for col, header in enumerate(map_headers, 1):
        cell = ws8.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    mappings = [
        ["id", "ID", "id", "id", "sku"],
        ["sku", "SKU", "reference", "title (variant)", "sku"],
        ["nombre_yolitia", "Name", "name (i18n)", "title", "name"],
        ["handle", "(auto)", "link_rewrite", "handle", "url_key"],
        ["descripcion_corta", "Short description", "description_short", "subtitle", "short_description"],
        ["descripcion_larga", "Description", "description", "description", "description"],
        ["precio", "Regular price", "price", "calculated_price", "price"],
        ["precio_comparacion", "Sale price", "reduction", "compare_at_price", "special_price"],
        ["categoria", "Categories", "id_category_default", "categories", "category_id"],
        ["seo.meta_keywords", "Tags", "meta_keywords", "tags", "meta_keyword"],
        ["imagenes[].url", "Images", "images", "images", "image"],
        ["imagenes[].alt_text", "(meta)", "legend (i18n)", "alt_text", "label"],
        ["inventario.cantidad", "Stock", "quantity", "inventory_quantity", "qty"],
        ["peso_gramos", "Weight", "weight", "weight", "weight"],
        ["material", "Attribute 1 value", "Feature", "material", "material"],
        ["activo + estado", "Published", "active", "status", "status"],
        ["visibilidad", "Visibility", "visibility", "-", "visibility"],
    ]
    
    for row_idx, mapping in enumerate(mappings, 2):
        for col, value in enumerate(mapping, 1):
            cell = ws8.cell(row=row_idx, column=col, value=value)
            cell.border = border
    
    wb.save(str(XLSX_FILE))
    print(f"  XLSX: {XLSX_FILE} (8 hojas)")


def main():
    print("=" * 60)
    print("BUILD DATABASE - YOLITIA")
    print("=" * 60)
    
    data = load_input()
    productos = data["productos"]
    
    print(f"\nProductos cargados: {len(productos)}")
    
    print("\nConstruyendo categorias...")
    categories = build_categories(productos)
    print(f"  Categorias: {len([c for c in categories.values() if c['parent_id'] is None])}")
    print(f"  Subcategorias: {len([c for c in categories.values() if c['parent_id'] is not None])}")
    
    print("\nGenerando archivos...")
    
    build_sqlite(productos, categories)
    save_schema_sql()
    build_json_universal(productos, categories)
    build_woocommerce_csv(productos, categories)
    build_prestashop_csv(productos, categories)
    build_medusa_json(productos, categories)
    build_xlsx(productos, categories)
    
    print(f"\n{'=' * 60}")
    print("ARCHIVOS GENERADOS:")
    print(f"{'=' * 60}")
    
    files = [
        ("SQLite DB", DB_FILE),
        ("SQL Schema", SCHEMA_FILE),
        ("JSON Universal", JSON_FILE),
        ("WooCommerce CSV", WOO_CSV),
        ("PrestaShop CSV", PRESTA_CSV),
        ("Medusa.js JSON", MEDUSA_JSON),
        ("XLSX", XLSX_FILE),
    ]
    
    for name, path in files:
        if path.exists():
            size = path.stat().st_size
            print(f"  [OK] {name:<20} {path.name:<40} {size:>10,} bytes")
        else:
            print(f"  [!!] {name:<20} {path.name:<40} {'MISSING':>10}")
    
    print(f"\n{'=' * 60}")


if __name__ == "__main__":
    main()
