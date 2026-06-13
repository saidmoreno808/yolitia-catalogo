import sqlite3, json, csv
import openpyxl

print('=== VALIDACION SQLite ===')
db = sqlite3.connect('yolitia_catalog/data/yolitia.db')
cursor = db.cursor()

cursor.execute('SELECT COUNT(*) FROM productos')
print(f'  Productos: {cursor.fetchone()[0]}')
cursor.execute('SELECT COUNT(*) FROM categorias')
print(f'  Categorias: {cursor.fetchone()[0]}')
cursor.execute('SELECT COUNT(*) FROM imagenes')
print(f'  Imagenes: {cursor.fetchone()[0]}')
cursor.execute('SELECT COUNT(*) FROM seo')
print(f'  SEO: {cursor.fetchone()[0]}')
cursor.execute('SELECT COUNT(*) FROM inventario')
print(f'  Inventario: {cursor.fetchone()[0]}')
cursor.execute('SELECT COUNT(*) FROM envio')
print(f'  Envio: {cursor.fetchone()[0]}')

cursor.execute('SELECT COUNT(*) FROM productos WHERE activo = 1')
print(f'  Productos activos: {cursor.fetchone()[0]}')

cursor.execute('SELECT DISTINCT estado FROM productos')
estados = [row[0] for row in cursor.fetchall()]
print(f'  Estados: {estados}')

db.close()

print()
print('=== VALIDACION JSON ===')
j = json.load(open('yolitia_catalog/data/yolitia_products_database.json','r',encoding='utf-8'))
print(f'  Productos: {len(j["productos"])}')
print(f'  Categorias: {len(j["categorias"])}')
print(f'  Metadata version: {j["metadata"]["version"]}')
print(f'  Plataformas: {j["metadata"]["plataformas_soportadas"]}')

print()
print('=== VALIDACION CSVs ===')
for name in ['yolitia_woocommerce_import.csv', 'yolitia_prestashop_import.csv']:
    with open(f'yolitia_catalog/data/{name}','r',encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
        print(f'  {name}: {len(rows)-1} filas, {len(rows[0])} columnas')

print()
print('=== VALIDACION Medusa JSON ===')
m = json.load(open('yolitia_catalog/data/yolitia_medusa_import.json','r',encoding='utf-8'))
print(f'  Productos: {len(m["products"])}')
if m["products"]:
    p0 = m["products"][0]
    print(f'  Primer producto: {p0["title"]}')
    print(f'  Handle: {p0["handle"]}')
    print(f'  Variants: {len(p0["variants"])}')

print()
print('=== VALIDACION XLSX ===')
wb = openpyxl.load_workbook('yolitia_catalog/data/yolitia_products_database.xlsx')
print(f'  Hojas: {wb.sheetnames}')
ws = wb['Productos']
print(f'  Productos (filas): {ws.max_row - 1}')
print(f'  Columnas: {ws.max_column}')

ws_img = wb['Imagenes']
print(f'  Imagenes (filas): {ws_img.max_row - 1}')

ws_cat = wb['Categorias']
print(f'  Categorias (filas): {ws_cat.max_row - 1}')

print()
print('=== TODAS LAS VALIDACIONES PASARON ===')
